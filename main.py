#pip install streamlit-option-menu
# Importando as bibliotecas necessárias

import streamlit as st
# Para trabalhar com datas
import datetime
# Pandas para o gerênciamento dos dados
import pandas as pd
# Numpy para trabalhar com arrays e funções matemáticas
import numpy as np
# Matplotlib e Seaborn para plotar gráficos 
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
import plotly.graph_objects as go
from plotly.offline import iplot
import altair as alt
from io import BytesIO
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
##
import matplotlib.dates as mdates
###
# Prophet para a predição de vendas
from prophet import Prophet
from prophet.plot import add_changepoints_to_plot
from prophet.plot import plot_plotly, plot_components_plotly
from streamlit_option_menu import option_menu
import mysql.connector
import base64
import io 
###
from neuralprophet import NeuralProphet
from statsmodels.tsa.arima.model import ARIMA
from pycaret.datasets import get_data
from pycaret.time_series import *


st.set_page_config(layout="wide")

def main():

    ####Desabilitar tela de login
    st.session_state['logged_in'] = True
    ####
    
    # Verificar se o usuário está logado
    if not is_user_logged_in():
        show_login_page()
    else:
        run_main_program()

def is_user_logged_in():
    # Verificar se o usuário está logado
    return st.session_state.get('logged_in', False)

def show_login_page():
    st.title("Login")
    username = st.text_input("Usuário")
    password = st.text_input("Senha", type="password")
    login_button = st.button("Login")

    if login_button:
        if username == "admin" and password == "ht7fty86":
            #st.success("Login successful!")
            st.session_state['logged_in'] = True
            st.experimental_rerun()  # Re-executar o aplicativo para atualizar a tela
        else:
            st.error("Invalid username or password")

@st.cache_data
def read_sheet(name, sheet):
        df = pd.read_excel(name, sheet, header=0)
        df = df.rename(columns={'QUANTIDADE': 'QUANT'})
        df = df.rename(columns={'DataInicioSemana': 'DATA'})
        
            
        df['Year'] = (df['DATA']//10000)
        df['Month'] = df['DATA'] - (df['Year'] * 10000)
        df['Month'] = df['Month'] // 100
        df['Day'] = df['DATA'] - ( df['Year'] * 10000 + df['Month'] * 100 )
        
        # Corrige algumas linhas que o dia é definido como sendo zero
        df.loc[df['Day'] < 1,'Day'] = 1
            
        # Cria um indice baseado no formato ano-semana
        df['AnoSemanaIdx'] =  df['Year'] * 100 + df['NumeroSemana']
        df = df.set_index('AnoSemanaIdx')
            
        # Corrige os dados das semanas repetidas
        df = df.groupby(df.index).agg({'NumeroSemana': 'first',
                                        'DATA': 'first', 
                                        'QUANT': 'sum',
                                        'Year': 'first',
                                        'Month': 'first',
                                        'Day': 'first'
                                        })
            
        
        # Coloca a data no formato DateTime em vez de inteiro
        df['DATA'] = pd.to_datetime(df[['Year','Month','Day']])
        
        # Corrige a semana extra no final do ano
        
        # 1 - Soma o valor da semana 53 na 52
        #fixedLastWeek = df.loc[df['NumeroSemana']==52, 'QUANT'].reset_index(drop=True) + df.loc[df['NumeroSemana']==53]['QUANT'].reset_index(drop=True)
        #cremoso[cremoso['NumeroSemana']==52]['QUANT']  = fixedLastWeek
        
        fixedLastWeek = df.loc[df['NumeroSemana']==52, 'QUANT'].reset_index(drop=True) + df.loc[df['NumeroSemana']==53]['QUANT'].reset_index(drop=True)
        dfFixedLastWeek = fixedLastWeek.to_frame()
        dfFixedLastWeek['anosemana'] = df[df['NumeroSemana']==52].index
        dfFixedLastWeek.set_index('anosemana', inplace=True)
        df.loc[df['NumeroSemana']==52,'QUANT'] = dfFixedLastWeek
            
        # 2 - Exclui a semana 53
        df = df.drop(df[df['NumeroSemana']==53].index)

        return df

@st.cache_data
def ler_nomes_das_planilhas(caminho_arquivo):
    workbook = openpyxl.load_workbook(caminho_arquivo)
        
    # Obtém os nomes das planilhas
    nomes_planilhas = workbook.sheetnames
        
    # Fecha o arquivo Excel
    workbook.close()
            
    return nomes_planilhas

@st.cache_data
def read_table(_connection, sheet):
    #df = pd.read_excel(name, sheet, header=0)
    cursor = _connection.cursor()
    cursor.execute(f"SELECT * FROM {sheet}")
    columns = [column[0] for column in cursor.description]
    data = cursor.fetchall()
    cursor.close()
    df = pd.DataFrame(data, columns=columns)
   
    df = df.rename(columns={'QUANTIDADE': 'QUANT'})
    df = df.rename(columns={'DataInicioSemana': 'DATA'})
            
    df['Year'] = (df['DATA']//10000)
    df['Month'] = df['DATA'] - (df['Year'] * 10000)
    df['Month'] = df['Month'] // 100
    df['Day'] = df['DATA'] - ( df['Year'] * 10000 + df['Month'] * 100 )
        
    # Corrige algumas linhas que o dia é definido como sendo zero
    df.loc[df['Day'] < 1,'Day'] = 1
            
    # Cria um indice baseado no formato ano-semana
    df['AnoSemanaIdx'] =  df['Year'] * 100 + df['NumeroSemana']
    df = df.set_index('AnoSemanaIdx')
            
    # Corrige os dados das semanas repetidas
    df = df.groupby(df.index).agg({'NumeroSemana': 'first',
                                    'DATA': 'first', 
                                    'QUANT': 'sum',
                                    'Year': 'first',
                                    'Month': 'first',
                                    'Day': 'first'
                                    })
            
        
    # Coloca a data no formato DateTime em vez de inteiro
    df['DATA'] = pd.to_datetime(df[['Year','Month','Day']])
        
    # Corrige a semana extra no final do ano
        
    # 1 - Soma o valor da semana 53 na 52
    #fixedLastWeek = df.loc[df['NumeroSemana']==52, 'QUANT'].reset_index(drop=True) + df.loc[df['NumeroSemana']==53]['QUANT'].reset_index(drop=True)
    #cremoso[cremoso['NumeroSemana']==52]['QUANT']  = fixedLastWeek
        
    fixedLastWeek = df.loc[df['NumeroSemana']==52, 'QUANT'].reset_index(drop=True) + df.loc[df['NumeroSemana']==53]['QUANT'].reset_index(drop=True)
    dfFixedLastWeek = fixedLastWeek.to_frame()
    dfFixedLastWeek['anosemana'] = df[df['NumeroSemana']==52].index
    dfFixedLastWeek.set_index('anosemana', inplace=True)
    df.loc[df['NumeroSemana']==52,'QUANT'] = dfFixedLastWeek
            
    # 2 - Exclui a semana 53
    df = df.drop(df[df['NumeroSemana']==53].index)
            
    return df
    
@st.cache_data
def get_table_names(connection):
    cursor = connection.cursor()
    cursor.execute("SHOW TABLES")
    tables = cursor.fetchall()
    table_names = [table[0] for table in tables]
    return table_names

@st.cache_data
def connect_to_mysql(user, password, host, database):
    try:
        connection = mysql.connector.connect(user=user, password=password,
                                            host=host,
                                            database=database)
        return connection
    except mysql.connector.Error as error:
        print("Erro ao conectar ao banco de dados MySQL:", error)
        return None

#@st.cache_data
def previsao(df, name, selected_graficos):       
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            #filtred_df.dropna()
            #filtred_df['QUANT'] = filtred_df['QUANT'].fillna(df['QUANT'][order - 1])
            return filtred_df
        
        @st.cache_data
        def create_profet_object(df, periodo):
            price_increase = pd.DataFrame({
              'holiday': 'price_increase',
              'ds': pd.to_datetime(['2018-01-01', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']),
              'lower_window': 0,
              'upper_window': 15,
            })
            
            specific_changepoints=['2018-01-05', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']
            
            prof = Prophet()
            
            
            prof.add_country_holidays(country_name='BR')
            df = df.rename(columns={'DATA': 'ds', 'QUANT': 'y'})
            prof.fit(df)
            future = prof.make_future_dataframe(periods=periodo, freq='W')
            #future.tail()
        
            forecast = prof.predict(future)
            #forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail()
            return (prof, forecast, df, future)
            
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
            
        @st.cache_data
        def generate_forecast_report(_prof, forecast, title):
            st.write(title)
            fig1 = prof.plot(forecast)
            st.pyplot(fig1)
            
        @st.cache_data
        def generate_components_report(_prof, forecast, title, df, future):
            prof = Prophet(weekly_seasonality=False)
            prof.add_seasonality(name='monthly', period=30.5, fourier_order=5)
            forecast = prof.fit(df).predict(future)
            fig2 = prof.plot_components(forecast)
            st.pyplot(fig2)
        
        @st.cache_data
        def generate_forecast_table(forecast, title, periodo):
            st.write(title)
            forecast_renamed = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(periodo)
            forecast_renamed = forecast_renamed.rename(columns={'ds': 'Data', 'yhat': 'Previsão média', 'yhat_lower': 'Previsão mínima', 'yhat_upper': 'Previsão máxima'})
            return forecast_renamed

        @st.cache_data
        def download_link_excel(df, filename):
            # Salvar o DataFrame em um arquivo Excel temporário
            temp_file = f"{filename}.xlsx"
            df.to_excel(temp_file, index=False)
            
            # Ler o arquivo Excel em bytes
            with open(temp_file, 'rb') as f:
                excel_data = f.read()
            
            # Excluir o arquivo temporário
            import os
            os.remove(temp_file)
            
            # Codificar em base64
            b64 = base64.b64encode(excel_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">Baixar tabela para o excel</a>'
            return href

        st.title('Previsão de demanda')
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro (semanas)", 1, 52)
            periodo = st.slider("Intervalo a ser previsto(semanas)", 1, 24, 12)
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        col1, col2 = st.columns(2)
        with col1:
            plot_historical_data(selected_product, selected_product_title)  
            
            
            prof, forecast, df, future = create_profet_object(selected_product, periodo)
            generate_forecast_report(prof, forecast, 'Previsão de demanda')

            forecast_table = generate_forecast_table(forecast, 'Previsão de Demanda - Tabela', periodo)
            st.write(forecast_table)
            ##########
            #forecast_table = generate_forecast_table(forecast, 'Previsão de Demanda - Tabela', periodo)
            #st.write(forecast_table)
            #######
    
        with col2:
            generate_components_report(prof, forecast, 'Componentes', df, future)
        
            fig = prof.plot_components(forecast)
                
            # Removendo o terceiro subplot que mostra os dias da semana
            #fig.delaxes(fig.axes[3])
            #fig.delaxes(fig.axes[2])
            #fig.delaxes(fig.axes[0])
            #st.pyplot(fig)
            
        st.markdown(download_link_excel(forecast_table, 'forecast'), unsafe_allow_html=True)

#@st.cache_data
def boxplot(df, name, selected_graficos):
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            return filtred_df
         
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #valor de y fixo
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
        @st.cache_data
        def boxplot_historical_data(df, title):
            st.write(title)
            fig, ax = plt.subplots()
            sns.boxplot(data=df, x='Year', y='QUANT', ax=ax)
            ax.set_title(title)
            ax.set_xlabel('Ano')
            ax.set_ylabel('Quantidade')
            plt.xticks(rotation=45)
            st.pyplot(fig)


        
        # Obtenção dos dados da interface do streamlit
        st.title("Boxplot")

        #primeiro gráfico
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro(semanas)", 1, 52)

        col1, col2 = st.columns(2)
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"

        with col1:
            plot_historical_data(selected_product, selected_product_title)  
        
        #segundo gráfico
        with col2:
            boxplot_historical_data(selected_product, selected_product_title2)

#@st.cache_data
def analise(df, name, selected_graficos):
        
            
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            return filtred_df
        
        @st.cache_data
        def create_profet_object(df, periodo):
            price_increase = pd.DataFrame({
              'holiday': 'price_increase',
              'ds': pd.to_datetime(['2018-01-01', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']),
              'lower_window': 0,
              'upper_window': 15,
            })
            
            specific_changepoints=['2018-01-05', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']
            
            prof = Prophet()
            
            
            prof.add_country_holidays(country_name='BR')
            df = df.rename(columns={'DATA': 'ds', 'QUANT': 'y'})
            prof.fit(df)
            future = prof.make_future_dataframe(periods=periodo, freq='W')
            forecast = prof.predict(future)
            return (prof, forecast, df, future)

        @st.cache_data
        def train_test_split(df):
            # Ordena o DataFrame pela data
            df_sorted = df.sort_values(by='DATA')
            # Calcula o índice para separar os dados de treinamento e teste
            split_index = int(len(df_sorted) * 0.95)  # porcentagem para treinamento 0.9 = 90%
            
            tamanhoPrevisao = int(len(df_sorted) * 0.05)
            
            train_df = df_sorted.iloc[:split_index]
            test_df = df_sorted.iloc[split_index:]
            return train_df, test_df, tamanhoPrevisao
            
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #valor de y fixo
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
        @st.cache_data
        def boxplot_historical_data(df, title):
            st.write(title)
            fig, ax = plt.subplots()
            sns.boxplot(data=df, x='Year', y='QUANT', ax=ax)
            ax.set_title(title)
            ax.set_xlabel('Ano')
            ax.set_ylabel('Quantidade')
            plt.xticks(rotation=45)
            st.pyplot(fig)
            
        @st.cache_data
        def generate_forecast_report(_prof, forecast, title):
            st.write(title)
            fig1 = _prof.plot(forecast)
            st.pyplot(fig1)

        @st.cache_data
        def generate_forecast_report2(_prof, forecast, title, tamanhoPrevisao):
            #tamanhoPrevisao += 1
            st.write(title)
            # Filtra as últimas num_semanas semanas
            forecast_filtered = forecast.tail(tamanhoPrevisao)
            fig1 = _prof.plot(forecast_filtered)
            st.pyplot(fig1)
            
            
        @st.cache_data
        def generate_components_report(_prof, forecast, title, df, future):
            _prof = Prophet(weekly_seasonality=False)
            _prof.add_seasonality(name='monthly', period=30.5, fourier_order=5)
            forecast = _prof.fit(df).predict(future)
            fig2 = _prof.plot_components(forecast)
            st.pyplot(fig2)
        
        @st.cache_data
        def ler_nomes_das_planilhas(caminho_arquivo):
            workbook = openpyxl.load_workbook(caminho_arquivo)
        
            # Obtém os nomes das planilhas
            nomes_planilhas = workbook.sheetnames
        
            # Fecha o arquivo Excel
            workbook.close()
            
            return nomes_planilhas
        
        @st.cache_data
        def generate_forecast_table(forecast, periodo):
            forecast_renamed = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(periodo)
            forecast_renamed = forecast_renamed.rename(columns={'ds': 'Data', 'yhat': 'Previsão média', 'yhat_lower': 'Previsão mínima', 'yhat_upper': 'Previsão máxima'})
            return forecast_renamed

        @st.cache_data
        def calcular_porcentagem_entre_min_max(prev_min, prev_max, test_data):
            total = len(test_data)
            dentro_intervalo = ((test_data >= prev_min) & (test_data <= prev_max)).sum()
            percentual = dentro_intervalo / total * 100
            return percentual

        @st.cache_data
        def calcular_erros(test_data, forecast_table, st, option2):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                # Calcular o Erro Médio Absoluto (MAE)
                mae = mean_absolute_error(test_data['QUANT'], forecast_table['Previsão média'])
                st.write("Erro Médio Absoluto (MAE):", int(mae))
            
            APE = mae / test_data['QUANT'] * 100
            
            # Calcular o Erro Percentual Absoluto Médio (MAPE)

            MAPE = APE.mean()
            
            # Configurar as cores para o gráfico de medidor
            if MAPE < 10:
                bar_color = "green"
            elif MAPE < 20:
                bar_color = "yellow"
            else:
                bar_color = "red"    
            
            # Criar o gráfico de medidor
            fig = go.Figure(go.Indicator(
                mode = "gauge+number",
                value = MAPE,
                title = {'text': "Erro Percentual Absoluto Médio (MAPE)"},
                gauge = {
                    'axis': {'range': [None, 100]},
                    'bar': {'color': bar_color, 'thickness': 1},  # Espessura máxima da barra
                    'bgcolor': 'black',
                    'borderwidth': 0,  # Remover bordas
                    'bordercolor': 'black',  # Ajustar a cor da borda para preto
                    'steps': [
                        {'range': [0, 100], 'color': 'gray'}  # Fundo preto
                    ],
                    'threshold': {
                        'line': {'color': "black", 'width': 4},
                        'thickness': 0.75,
                        'value': MAPE
                    }
                }
            ))
            with col2: 
                st.write("Erro Percentual Absoluto Médio (MAPE):", round(MAPE, 2), "%")
                #st.plotly_chart(fig)
            with col3:
                # Calcular o Root Mean Square Error (RMSE)
                RMSE = np.sqrt(mean_squared_error(test_data['QUANT'], forecast_table['Previsão média']))
                st.write("Raiz do Erro Quadrático Médio (RMSE):", int(RMSE))
            with col4:
                # Frequência que sai dos limites
                percentual_min_max = calcular_porcentagem_entre_min_max(forecast_table['Previsão mínima'], forecast_table['Previsão máxima'], test_data['QUANT'])
                st.write("Frequência entre o intervalo:", round(percentual_min_max, 2), "%")
            
            
        
            # Calcular os erros
            errors = test_data['QUANT'] - forecast_table['Previsão média']
        
            if option2:
                st.write("Diferenças entre os valores e as previsões:")
                st.write(errors)

            col1, col2 = st.columns(2)
              
            # Plotar o histograma dos erros
            plt.figure(figsize=(8, 6))
            plt.hist(errors, bins=6, color='skyblue', edgecolor='black')
            plt.title('Histograma dos Erros')
            plt.xlabel('Erro')
            plt.ylabel('Frequência')
            plt.grid(True)
            with col1: 
                st.pyplot(plt)
            
            # Erros em porcentagem
            errors_percent = (test_data['QUANT'] - forecast_table['Previsão média']) / test_data['QUANT'] * 100
            if option2:
                st.write("Diferenças entre os valores e as previsões em porcentagem")
                st.write(errors_percent)
            
            # Plotar o histograma dos erros percentuais
            plt.figure(figsize=(8, 6))
            plt.hist(errors_percent, bins=6, color='skyblue', edgecolor='black')
            plt.title('Histograma dos Erros percentuais')
            plt.xlabel('Erro %')
            plt.ylabel('Frequência')
            plt.grid(True)

            with col2: 
                st.pyplot(plt)


        # Obtenção dos dados da interface do streamlit
        st.title("Medidas de análise de precisão")
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro(semanas)", 1, 52)

        col1, col2 = st.columns(2)

        ##grafico sem média móvel
        selected_productS = df
        selected_productS = moving_average_filter(selected_productS, 1)
        selected_product_titleS = selected_graficos + " - Ordem do Filtro: " + str(1) + " (semanas)"
        with col1:
            plot_historical_data(selected_productS, selected_product_titleS) 

        
        #primeiro gráfico
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        with col2:
            plot_historical_data(selected_product, selected_product_title) 
        
        
        #segundo gráfico
        #boxplot_historical_data(selected_product, selected_product_title2)

        #dividindo dados para treino e teste
        train_data, test_data, tamanhoPrevisao = train_test_split(selected_product)
        tamanhoPrevisao += 1
        
        #terceiro gráfico/previsão
        #periodo = st.slider("Intervalo a ser previsto(semanas)", 12, 38)      
        
        prof_train, forecast_train, df_train, future_train = create_profet_object(train_data, tamanhoPrevisao)
        #prof_train, forecast_train, df_train, future_train = create_profet_object(train_data, 38)
    
        
        col1, col2, col3 = st.columns(3)
        with col1:
            generate_forecast_report(prof_train, forecast_train, 'Previsão de demanda')
        with col2:     
            prof_test, forecast_test, df_test, future_test = create_profet_object(test_data, 0)
            generate_forecast_report2(prof_test, forecast_train, 'Previsão vs valor', tamanhoPrevisao)
                     
            ##tabela previsão
            forecast_table = generate_forecast_table(forecast_train, tamanhoPrevisao)
            #forecast_table = generate_forecast_table(forecast_train, 'Previsão de Demanda - Tabela', 38)
        with col3:
            #dividindo dados para treino e teste
            train_dataS, test_dataS, tamanhoPrevisao = train_test_split(selected_productS)
            prof_testS, forecast_testS, df_testS, future_testS = create_profet_object(test_dataS, 0)
            generate_forecast_report2(prof_testS, forecast_train, 'VS Previsão sem média móvel', tamanhoPrevisao)           
                
           

        option2 = st.checkbox('Mostrar detalhes dos gráficos')
        #option2 = 'Mostrar detalhes dos gráficos'
        #tabela previsão
        if option2:
            st.write('Previsão de Demanda - Tabela')
            st.write(forecast_table)
        

        # Criar uma caixa de seleção
        #option = st.checkbox('Mostrar dados de análise de precisão')
        option = 'Mostrar dados de análise de precisão'
    
        
        # Verificar se a caixa de seleção está marcada
        
        if option:
            #if periodo != 0:
            #    forecast_table = forecast_table[:-periodo]
            #    test_data = test_data[:-periodo]
    
            # Redefinir os índices para garantir que estejam alinhados corretamente
            test_data.reset_index(drop=True, inplace=True)
            forecast_table.reset_index(drop=True, inplace=True)
            
            if option2:
                st.write("Dados para teste:")
                st.write(test_data)     
                st.write("Coluna previsão média")
                st.write(forecast_table['Previsão média'])
                st.write("Coluna dados para teste")
                st.write(test_data['QUANT'])

            option3 = st.checkbox('Mostrar em relação aos dados sem média móvel')
            if option3:
                ############ Em relação a sem média móvel
                st.write('Em relação ao dado sem média móvel')
                selected_product = df
                selected_product = moving_average_filter(selected_product, 1)
                selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(1) + " (semanas)"
                train_data, test_data, tamanhoPrevisao = train_test_split(selected_product)
                test_data.reset_index(drop=True, inplace=True)
                
                if option2:
                    st.write("Dados para teste-sem média móvel:")
                    st.write(test_data)
                
            ####função
            calcular_erros(test_data, forecast_table, st, option2)

#@st.cache_data
def previsaoNeural(df, name, selected_graficos):       
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            #filtred_df.dropna()
            #filtred_df['QUANT'] = filtred_df['QUANT'].fillna(df['QUANT'][order - 1])
            return filtred_df
        
        @st.cache_data
        def create_profet_object(df, periodo):      
            price_increase = pd.DataFrame({
              'holiday': 'price_increase',
              'ds': pd.to_datetime(['2018-01-01', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']),
              'lower_window': 0,
              'upper_window': 15,
            })
            
            specific_changepoints=['2018-01-05', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']
            
            #st.write(df.head())
            
            df = df.rename(columns={'DATA': 'ds', 'QUANT': 'y'})
            duplicate_ds = df[df.duplicated(subset=['ds'], keep=False)]
            #if not duplicate_ds.empty:
                #st.write("Duplicatas encontradas na coluna 'ds', só a primeira foi mantida:")
                #st.write(duplicate_ds)
            df = df.drop_duplicates(subset=['ds'], keep='first')
            # Seleção das colunas 'ds' e 'y'
            df = df.loc[:, ['ds', 'y']]
            #df = df.drop(columns=['AnoSemanaIdx'])
            


            # Criação do objeto NeuralProphet
            prof = NeuralProphet()
            
            # Ajuste do modelo com os dados históricos
            prof.fit(df)

            #future = prof.make_future_dataframe(periods=periodo)
            future = prof.make_future_dataframe(df=df, periods=periodo)
            
            #future = pd.DataFrame({
            #    'ds': pd.date_range(start=df['ds'].max(), periods=periodo, freq='W')
            #})
            
            # Previsão dos valores futuros
            forecast = prof.predict(future)

            ###########
            # Configurações para plotagem
            fig, ax = plt.subplots()
            ax.plot(df['ds'], df['y'], label='Histórico')
            ax.plot(forecast['ds'], forecast['yhat1'], label='Previsão', linestyle='--')
    
            # Formatação do eixo x para exibir datas
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    
            # Adiciona legenda e rótulos aos eixos
            ax.legend()
            ax.set_title('Previsão com NeuralProphet')
            ax.set_xlabel('Data')
            ax.set_ylabel('Valor')
            ax.grid(True)
    
            # Exibe o gráfico na interface do Streamlit
            #st.pyplot(fig)
            #############
            
            return (prof, forecast, df, future, fig)
            
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
            
        @st.cache_data
        def generate_forecast_report(_prof, forecast, title):
            st.write(title)
            fig1 = prof.plot(forecast)
            st.pyplot(fig1)
            
        @st.cache_data
        def generate_components_report(_prof, forecast, title, df, future):
            prof = Prophet(weekly_seasonality=False)
            prof.add_seasonality(name='monthly', period=30.5, fourier_order=5)
            forecast = prof.fit(df).predict(future)
            fig2 = prof.plot_components(forecast)
            st.pyplot(fig2)
        
        @st.cache_data
        def generate_forecast_table(forecast, title, periodo):
            st.write(title)
            forecast_renamed = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(periodo)
            forecast_renamed = forecast_renamed.rename(columns={'ds': 'Data', 'yhat': 'Previsão média', 'yhat_lower': 'Previsão mínima', 'yhat_upper': 'Previsão máxima'})
            return forecast_renamed
            
        @st.cache_data
        def download_link_excel(df, filename):
            # Salvar o DataFrame em um arquivo Excel temporário
            temp_file = f"{filename}.xlsx"
            df.to_excel(temp_file, index=False)
            
            # Ler o arquivo Excel em bytes
            with open(temp_file, 'rb') as f:
                excel_data = f.read()
            
            # Excluir o arquivo temporário
            import os
            os.remove(temp_file)
            
            # Codificar em base64
            b64 = base64.b64encode(excel_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">Baixar tabela para o excel</a>'
            return href

        st.title('Previsão de demanda')
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro (semanas)", 1, 52)
            periodo = st.slider("Intervalo a ser previsto(semanas)", 1, 24, 12) + 1
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        col1, col2 = st.columns(2)
        with col1:
            plot_historical_data(selected_product, selected_product_title)  
            
            
            prof, forecast, df, future, fig = create_profet_object(selected_product, periodo)
            st.write(fig)
            #generate_forecast_report(prof, forecast, 'Previsão de demanda')
            
            ##########
            #forecast_table = generate_forecast_table(forecast, 'Previsão de Demanda - Tabela', periodo)
            #st.write(forecast_table)
            #######
    
        with col2:
            generate_components_report(prof, forecast, 'Componentes', df, future)
        
            #fig = prof.plot_components(forecast)
                
            # Removendo o terceiro subplot que mostra os dias da semana
            #fig.delaxes(fig.axes[3])
            #fig.delaxes(fig.axes[2])
            #fig.delaxes(fig.axes[0])
            #st.pyplot(fig)
            
        #forecast_table = generate_forecast_table(forecast, 'Previsão de Demanda - Tabela', periodo)
        forecast_renamed = forecast.rename(columns={'ds': 'Data', 'yhat1': 'Previsão média'})
        forecast_renamed = forecast_renamed.drop(columns=['y', 'trend', 'season_weekly', 'season_yearly'])
        forecast_renamed = forecast_renamed.iloc[1:]
        st.write(forecast_renamed)
        st.markdown(download_link_excel(forecast_renamed, 'forecast'), unsafe_allow_html=True)

#@st.cache_data
def analiseNeural(df, name, selected_graficos):           
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            return filtred_df
        
        @st.cache_data
        def create_profet_object(df, periodo):
            price_increase = pd.DataFrame({
              'holiday': 'price_increase',
              'ds': pd.to_datetime(['2018-01-01', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']),
              'lower_window': 0,
              'upper_window': 15,
            })
            
            specific_changepoints=['2018-01-05', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']
            
            #st.write(df.head())
            
            df = df.rename(columns={'DATA': 'ds', 'QUANT': 'y'})
            duplicate_ds = df[df.duplicated(subset=['ds'], keep=False)]
            #if not duplicate_ds.empty:
                #st.write("Duplicatas encontradas na coluna 'ds', só a primeira foi mantida:")
                #st.write(duplicate_ds)
            df = df.drop_duplicates(subset=['ds'], keep='first')
            # Seleção das colunas 'ds' e 'y'
            df = df.loc[:, ['ds', 'y']]
            #df = df.drop(columns=['AnoSemanaIdx'])



            # Criação do objeto NeuralProphet
            prof = NeuralProphet()
            
            # Ajuste do modelo com os dados históricos
            prof.fit(df)

            #future = prof.make_future_dataframe(periods=periodo)
            future = prof.make_future_dataframe(df=df, periods=periodo)
            
            #future = pd.DataFrame({
            #    'ds': pd.date_range(start=df['ds'].max(), periods=periodo, freq='W')
            #})
            
            # Previsão dos valores futuros
            forecast = prof.predict(future)

            ###########
            # Configurações para plotagem
            fig, ax = plt.subplots()
            ax.plot(df['ds'], df['y'], label='Histórico')
            ax.plot(forecast['ds'], forecast['yhat1'], label='Previsão', linestyle='--')
    
            # Formatação do eixo x para exibir datas
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    
            # Adiciona legenda e rótulos aos eixos
            ax.legend()
            ax.set_title('Previsão com NeuralProphet')
            ax.set_xlabel('Data')
            ax.set_ylabel('Valor')
            ax.grid(True)
    
            # Exibe o gráfico na interface do Streamlit
            #st.pyplot(fig)
            #############
            
            return (prof, forecast, df, future, fig)

        @st.cache_data
        def train_test_split(df):
            # Ordena o DataFrame pela data
            df_sorted = df.sort_values(by='DATA')
            # Calcula o índice para separar os dados de treinamento e teste
            split_index = int(len(df_sorted) * 0.95)  # porcentagem para treinamento 0.9 = 90%
            
            tamanhoPrevisao = int(len(df_sorted) * 0.05)
            
            train_df = df_sorted.iloc[:split_index]
            test_df = df_sorted.iloc[split_index:]
            
            return train_df, test_df, tamanhoPrevisao
            
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #valor de y fixo
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
        @st.cache_data
        def boxplot_historical_data(df, title):
            st.write(title)
            fig, ax = plt.subplots()
            sns.boxplot(data=df, x='Year', y='QUANT', ax=ax)
            ax.set_title(title)
            ax.set_xlabel('Ano')
            ax.set_ylabel('Quantidade')
            plt.xticks(rotation=45)
            st.pyplot(fig)
            
        @st.cache_data
        def generate_forecast_report(_prof, forecast, title):
            st.write(title)
            fig1 = _prof.plot(forecast)
            st.pyplot(fig1)

        @st.cache_data
        def generate_forecast_report2(_prof, forecast, title, tamanhoPrevisao):
            #tamanhoPrevisao += 1
            st.write(title)
            # Filtra as últimas num_semanas semanas
            forecast_filtered = forecast.tail(tamanhoPrevisao)
            fig1 = _prof.plot(forecast_filtered)
            st.pyplot(fig1)
            
            
        @st.cache_data
        def generate_components_report(_prof, forecast, title, df, future):
            _prof = Prophet(weekly_seasonality=False)
            _prof.add_seasonality(name='monthly', period=30.5, fourier_order=5)
            forecast = _prof.fit(df).predict(future)
            fig2 = _prof.plot_components(forecast)
            st.pyplot(fig2)
        
        @st.cache_data
        def ler_nomes_das_planilhas(caminho_arquivo):
            workbook = openpyxl.load_workbook(caminho_arquivo)
        
            # Obtém os nomes das planilhas
            nomes_planilhas = workbook.sheetnames
        
            # Fecha o arquivo Excel
            workbook.close()
            
            return nomes_planilhas
        
        @st.cache_data
        def generate_forecast_table(forecast, periodo):
            forecast_renamed = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(periodo)
            forecast_renamed = forecast_renamed.rename(columns={'ds': 'Data', 'yhat': 'Previsão média', 'yhat_lower': 'Previsão mínima', 'yhat_upper': 'Previsão máxima'})
            return forecast_renamed

        @st.cache_data
        def calcular_porcentagem_entre_min_max(prev_min, prev_max, test_data):
            total = len(test_data)
            dentro_intervalo = ((test_data >= prev_min) & (test_data <= prev_max)).sum()
            percentual = dentro_intervalo / total * 100
            return percentual

        @st.cache_data
        def calcular_erros(test_data, forecast_train, st, option2):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                # Calcular o Erro Médio Absoluto (MAE)
                mae = mean_absolute_error(test_data['QUANT'], forecast_train['yhat1'])
                st.write("Erro Médio Absoluto (MAE):", int(mae))
            
            APE = mae / test_data['QUANT'] * 100
            
            # Calcular o Erro Percentual Absoluto Médio (MAPE)

            MAPE = APE.mean()
            
            # Configurar as cores para o gráfico de medidor
            if MAPE < 10:
                bar_color = "green"
            elif MAPE < 20:
                bar_color = "yellow"
            else:
                bar_color = "red"    
            
            # Criar o gráfico de medidor
            fig = go.Figure(go.Indicator(
                mode = "gauge+number",
                value = MAPE,
                title = {'text': "Erro Percentual Absoluto Médio (MAPE)"},
                gauge = {
                    'axis': {'range': [None, 100]},
                    'bar': {'color': bar_color, 'thickness': 1},  # Espessura máxima da barra
                    'bgcolor': 'black',
                    'borderwidth': 0,  # Remover bordas
                    'bordercolor': 'black',  # Ajustar a cor da borda para preto
                    'steps': [
                        {'range': [0, 100], 'color': 'gray'}  # Fundo preto
                    ],
                    'threshold': {
                        'line': {'color': "black", 'width': 4},
                        'thickness': 0.75,
                        'value': MAPE
                    }
                }
            ))
            with col2: 
                st.write("Erro Percentual Absoluto Médio (MAPE):", round(MAPE, 2), "%")
                #st.plotly_chart(fig)
            with col3:
                # Calcular o Root Mean Square Error (RMSE)
                RMSE = np.sqrt(mean_squared_error(test_data['QUANT'], forecast_train['yhat1']))
                st.write("Raiz do Erro Quadrático Médio (RMSE):", int(RMSE))

            
            
        
            # Calcular os erros
            errors = test_data['QUANT'] - forecast_train['yhat1']
        
            if option2:
                st.write("Diferenças entre os valores e as previsões:")
                st.write(errors)

            col1, col2 = st.columns(2)
              
            # Plotar o histograma dos erros
            plt.figure(figsize=(8, 6))
            plt.hist(errors, bins=6, color='skyblue', edgecolor='black')
            plt.title('Histograma dos Erros')
            plt.xlabel('Erro')
            plt.ylabel('Frequência')
            plt.grid(True)
            with col1: 
                st.pyplot(plt)
            
            # Erros em porcentagem
            errors_percent = (test_data['QUANT'] - forecast_train['yhat1']) / test_data['QUANT'] * 100
            if option2:
                st.write("Diferenças entre os valores e as previsões em porcentagem")
                st.write(errors_percent)
            
            # Plotar o histograma dos erros percentuais
            plt.figure(figsize=(8, 6))
            plt.hist(errors_percent, bins=6, color='skyblue', edgecolor='black')
            plt.title('Histograma dos Erros percentuais')
            plt.xlabel('Erro %')
            plt.ylabel('Frequência')
            plt.grid(True)

            with col2: 
                st.pyplot(plt)


        # Obtenção dos dados da interface do streamlit
        st.title("Medidas de análise de precisão")
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro(semanas)", 1, 52)

        col1, col2 = st.columns(2)

        ##grafico sem média móvel
        selected_productS = df
        selected_productS = moving_average_filter(selected_productS, 1)
        selected_product_titleS = selected_graficos + " - Ordem do Filtro: " + str(1) + " (semanas)"
        with col1:
            plot_historical_data(selected_productS, selected_product_titleS) 

        
        #primeiro gráfico
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        with col2:
            plot_historical_data(selected_product, selected_product_title) 
        
        
        #segundo gráfico
        #boxplot_historical_data(selected_product, selected_product_title2)

        #dividindo dados para treino e teste
        train_data, test_data, tamanhoPrevisao = train_test_split(selected_product)
        tamanhoPrevisao += 1
        tamanhoPrevisao += 1
    
        
        #terceiro gráfico/previsão
        #periodo = st.slider("Intervalo a ser previsto(semanas)", 12, 38)      
        
        prof_train, forecast_train, df_train, future_train, fig = create_profet_object(train_data, tamanhoPrevisao)
        #prof_train, forecast_train, df_train, future_train = create_profet_object(train_data, 38)
        forecast_train = forecast_train.iloc[1:]

        
        
        col1, col2, col3 = st.columns(3)
        with col1:
            #generate_forecast_report(prof_train, forecast_train, 'Previsão de demanda')
            st.pyplot(fig)
        with col2:     
            prof_test, forecast_test, df_test, future_test, fig = create_profet_object(test_data, 1)

            # Criar o gráfico
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Plotar a linha de previsão
            ax.plot(test_data['DATA'], forecast_train['yhat1'], color='orange', linestyle='--', label='Previsão')
            
            # Plotar a linha de dados reais
            ax.plot(test_data['DATA'], test_data['QUANT'], color='blue', label='Real')
            
            # Adicionar títulos e legendas
            ax.set_title('Previsão vs Real')
            ax.set_xlabel('Data')
            ax.set_ylabel('Quantidade')
            ax.legend()
            
            # Exibir o gráfico no Streamlit
            st.pyplot(fig)

            
            
        with col3:
            #dividindo dados para treino e teste
            train_dataS, test_dataS, tamanhoPrevisao = train_test_split(selected_productS)
            prof_testS, forecast_testS, df_testS, future_testS, fig = create_profet_object(test_dataS, 1)

            prof_test, forecast_test, df_test, future_test, fig = create_profet_object(test_data, 1)

            # Criar o gráfico
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Plotar a linha de previsão
            ax.plot(test_dataS['DATA'], forecast_train['yhat1'], color='orange', linestyle='--', label='Previsão')
            
            # Plotar a linha de dados reais
            ax.plot(test_dataS['DATA'], test_dataS['QUANT'], color='blue', label='Real')
            
            # Adicionar títulos e legendas
            ax.set_title('Previsão vs Real (sem média móvel)')
            ax.set_xlabel('Data')
            ax.set_ylabel('Quantidade')
            ax.legend()
            
            # Exibir o gráfico no Streamlit
            st.pyplot(fig)
                
           
        option2 = st.checkbox('Mostrar detalhes dos gráficos')
        #option2 = 'Mostrar detalhes dos gráficos'
        #tabela previsão
        if option2:
            st.write('Previsão de Demanda - Tabela')
            st.write(forecast_train)

        
    
        # Criar uma caixa de seleção
        #option = st.checkbox('Mostrar dados de análise de precisão')
        option = 'Mostrar dados de análise de precisão'
    
        
        # Verificar se a caixa de seleção está marcada
        
        if option:
            #if periodo != 0:
            #    forecast_table = forecast_table[:-periodo]
            #    test_data = test_data[:-periodo]
    
            # Redefinir os índices para garantir que estejam alinhados corretamente
            test_data.reset_index(drop=True, inplace=True)
            forecast_train.reset_index(drop=True, inplace=True)
            
            if option2:
                st.write("Dados para teste:")
                st.write(test_data)     
                st.write("Coluna previsão média")
                st.write(forecast_train['yhat1'])
                st.write("Coluna dados para teste")
                st.write(test_data['QUANT'])
                #test_data['DATA'] para o eixo x, e plotar forecast_train['yhat1'] como uma linha azul com legenda de previsão, e no mesmo gráfico plotar test_data['QUANT']

            option3 = st.checkbox('Mostrar em relação aos dados sem média móvel')
            if option3:
                ############ Em relação a sem média móvel
                st.write('Em relação ao dado sem média móvel')
                selected_product = df
                selected_product = moving_average_filter(selected_product, 1)
                selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(1) + " (semanas)"
                train_data, test_data, tamanhoPrevisao = train_test_split(selected_product)
                test_data.reset_index(drop=True, inplace=True)
                
                if option2:
                    st.write("Dados para teste-sem média móvel:")
                    st.write(test_data)
                
            ####função
            calcular_erros(test_data, forecast_train, st, option2)

#@st.cache_data
def previsaoArima(df, name, selected_graficos):       
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            #filtred_df.dropna()
            #filtred_df['QUANT'] = filtred_df['QUANT'].fillna(df['QUANT'][order - 1])
            return filtred_df
        
        @st.cache_data
        def create_profet_object(df, periodo, p, d, q):
            price_increase = pd.DataFrame({
              'holiday': 'price_increase',
              'ds': pd.to_datetime(['2018-01-01', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']),
              'lower_window': 0,
              'upper_window': 15,
            })
            
            specific_changepoints=['2018-01-05', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']
            
            #st.write(df.head())
            
            df = df.rename(columns={'DATA': 'ds', 'QUANT': 'y'})
            duplicate_ds = df[df.duplicated(subset=['ds'], keep=False)]
            #if not duplicate_ds.empty:
                #st.write("Duplicatas encontradas na coluna 'ds', só a primeira foi mantida:")
                #st.write(duplicate_ds)
            df = df.drop_duplicates(subset=['ds'], keep='first')
            # Seleção das colunas 'ds' e 'y'
            df = df.loc[:, ['ds', 'y']]
            #df = df.drop(columns=['AnoSemanaIdx'])


            #p = st.slider('Parâmetro p (AutoRegressivo)', 0, 5, 1)     
            #d = st.slider('Parâmetro d (Integração)', 0, 2, 1)

            
            # Ajusta o modelo ARIMA
            prof = ARIMA(df['y'], order=(p, d, q))  # Componente MA definido como 0
            prof_fit = prof.fit()
        
            # Cria um DataFrame com datas futuras para previsão
            future_dates = pd.date_range(df['ds'].iloc[-1] + pd.Timedelta(weeks=1), periods=periodo, freq='W')
            forecast = prof_fit.forecast(steps=periodo)
            future = pd.DataFrame({'ds': future_dates, 'y': forecast})

        
            # Configurações para plotagem
            fig, ax = plt.subplots()
            ax.plot(df['ds'], df['y'], label='Histórico')
            ax.plot(future['ds'], future['y'], label='Previsão', linestyle='--')
        
            # Formatação do eixo x para exibir datas
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
        
            # Adiciona legenda e rótulos aos eixos
            ax.legend()
            ax.set_title('Previsão com ARIMA')
            ax.set_xlabel('Data')
            ax.set_ylabel('Valor')
            ax.grid(True)
        
            return prof, forecast, df, future, fig, future_dates
            
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
            
        @st.cache_data
        def generate_forecast_report(_prof, forecast, title):
            st.write(title)
            fig1 = prof.plot(forecast)
            st.pyplot(fig1)
            
        @st.cache_data
        def generate_components_report(_prof, forecast, title, df, future):
            prof = Prophet(weekly_seasonality=False)
            prof.add_seasonality(name='monthly', period=30.5, fourier_order=5)
            forecast = prof.fit(df).predict(future)
            fig2 = prof.plot_components(forecast)
            st.pyplot(fig2)
        
        @st.cache_data
        def generate_forecast_table(forecast, title, periodo):
            st.write(title)
            #st.write(forecast)
            forecast_renamed = forecast[['ds', 'y']].tail(periodo)
            forecast_renamed = forecast_renamed.rename(columns={'ds': 'Data', 'y': 'Previsão média'})
            return forecast_renamed
            
        @st.cache_data
        def download_link_excel(df, filename):
            # Salvar o DataFrame em um arquivo Excel temporário
            temp_file = f"{filename}.xlsx"
            df.to_excel(temp_file, index=False)
            
            # Ler o arquivo Excel em bytes
            with open(temp_file, 'rb') as f:
                excel_data = f.read()
            
            # Excluir o arquivo temporário
            import os
            os.remove(temp_file)
            
            # Codificar em base64
            b64 = base64.b64encode(excel_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">Baixar tabela para o excel</a>'
            return href

        st.title('Previsão de demanda')
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro (semanas)", 1, 52)
            periodo = st.slider("Intervalo a ser previsto(semanas)", 1, 24, 12)
            p = st.slider('Parâmetro p (AutoRegressivo)', 0, 20, 1)     
            d = st.slider('Parâmetro d (Integração)', 0, 2, 1)
            q = st.slider('Parâmetro q (Média móvel)', 0, 25, 1)
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        col1, col2 = st.columns(2)
        with col1:
            plot_historical_data(selected_product, selected_product_title)  
            
            #generate_forecast_report(prof, forecast, 'Previsão de demanda')
            
            ##########
            #forecast_table = generate_forecast_table(forecast, 'Previsão de Demanda - Tabela', periodo)
            #st.write(forecast_table)
            #######
    
        with col2:
            prof, forecast, df, future, fig, future_dates = create_profet_object(selected_product, periodo, p, d, q)
            st.write(fig)
            #generate_components_report(prof, forecast, 'Componentes', df, future)
        
            #fig = prof.plot_components(forecast)
                
            # Removendo o terceiro subplot que mostra os dias da semana
            #fig.delaxes(fig.axes[3])
            #fig.delaxes(fig.axes[2])
            #fig.delaxes(fig.axes[0])
            #st.pyplot(fig)
        forecast = pd.DataFrame({'ds': future_dates, 'y': forecast})
        forecast_renamed = forecast.rename(columns={'ds': 'Data', 'y': 'Previsão média'})
        forecast_table = generate_forecast_table(forecast, 'Previsão de Demanda - Tabela', periodo)
        st.write(forecast_renamed)

        st.markdown(download_link_excel(forecast_renamed, 'forecast'), unsafe_allow_html=True)
    
#@st.cache_data
def analiseArima(df, name, selected_graficos):            
        @st.cache_data
        def moving_average_filter(df, order):
            filtred_df = df.copy()
            filtred_df['QUANT'] = filtred_df['QUANT'].rolling(order).mean()
            return filtred_df
        
        @st.cache_data
        def create_profet_object(df, periodo, p, d, q):
            price_increase = pd.DataFrame({
              'holiday': 'price_increase',
              'ds': pd.to_datetime(['2018-01-01', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']),
              'lower_window': 0,
              'upper_window': 15,
            })
            
            specific_changepoints=['2018-01-05', '2018-04-01', '2018-11-01',
                                    '2019-09-01', '2020-03-01', '2020-06-01',
                                    '2021-02-01', '2021-06-01', '2022-02-01',
                                    '2022-05-01', '2022-08-01', '2023-01-01']
            
            #st.write(df.head())
            
            df = df.rename(columns={'DATA': 'ds', 'QUANT': 'y'})
            duplicate_ds = df[df.duplicated(subset=['ds'], keep=False)]
            #if not duplicate_ds.empty:
                #st.write("Duplicatas encontradas na coluna 'ds', só a primeira foi mantida:")
                #st.write(duplicate_ds)
            df = df.drop_duplicates(subset=['ds'], keep='first')
            # Seleção das colunas 'ds' e 'y'
            df = df.loc[:, ['ds', 'y']]
            #df = df.drop(columns=['AnoSemanaIdx'])


            #p = st.slider('Parâmetro p (AutoRegressivo)', 0, 5, 1)     
            #d = st.slider('Parâmetro d (Integração)', 0, 2, 1)

            
            # Ajusta o modelo ARIMA
            prof = ARIMA(df['y'], order=(p, d, q))  # Componente MA definido como 0
            prof_fit = prof.fit()
        
            # Cria um DataFrame com datas futuras para previsão
            future_dates = pd.date_range(df['ds'].iloc[-1] + pd.Timedelta(weeks=1), periods=periodo, freq='W')
            forecast = prof_fit.forecast(steps=periodo)
            future = pd.DataFrame({'ds': future_dates, 'y': forecast})

        
            # Configurações para plotagem
            fig, ax = plt.subplots()
            ax.plot(df['ds'], df['y'], label='Histórico')
            ax.plot(future['ds'], future['y'], label='Previsão', linestyle='--')
        
            # Formatação do eixo x para exibir datas
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
        
            # Adiciona legenda e rótulos aos eixos
            ax.legend()
            ax.set_title('Previsão com ARIMA')
            ax.set_xlabel('Data')
            ax.set_ylabel('Valor')
            ax.grid(True)
        
            return prof, forecast, df, future, fig, future_dates

        @st.cache_data
        def train_test_split(df):
            # Ordena o DataFrame pela data
            df_sorted = df.sort_values(by='DATA')
            # Calcula o índice para separar os dados de treinamento e teste
            split_index = int(len(df_sorted) * 0.95)  # porcentagem para treinamento 0.9 = 90%
            
            tamanhoPrevisao = int(len(df_sorted) * 0.05)
            
            train_df = df_sorted.iloc[:split_index]
            test_df = df_sorted.iloc[split_index:]
            
            return train_df, test_df, tamanhoPrevisao
            
        @st.cache_data    
        def plot_historical_data(df, title):
            st.write(title)
            #st.line_chart(df.set_index('DATA'))
            chart = alt.Chart(df.reset_index()).mark_line().encode(
                x='DATA',
                y='QUANT'
                #valor de y fixo
                #y=alt.Y('QUANT', scale=alt.Scale(domain=[0, 200000]))
            ).interactive()
            st.altair_chart(chart, use_container_width=True)
            
        @st.cache_data
        def boxplot_historical_data(df, title):
            st.write(title)
            fig, ax = plt.subplots()
            sns.boxplot(data=df, x='Year', y='QUANT', ax=ax)
            ax.set_title(title)
            ax.set_xlabel('Ano')
            ax.set_ylabel('Quantidade')
            plt.xticks(rotation=45)
            st.pyplot(fig)
            
        @st.cache_data
        def generate_forecast_report(_prof, forecast, title):
            st.write(title)
            fig1 = _prof.plot(forecast)
            st.pyplot(fig1)

        @st.cache_data
        def generate_forecast_report2(_prof, forecast, title, tamanhoPrevisao):
            #tamanhoPrevisao += 1
            st.write(title)
            # Filtra as últimas num_semanas semanas
            forecast_filtered = forecast.tail(tamanhoPrevisao)
            fig1 = _prof.plot(forecast_filtered)
            st.pyplot(fig1)
            
            
        @st.cache_data
        def generate_components_report(_prof, forecast, title, df, future):
            _prof = Prophet(weekly_seasonality=False)
            _prof.add_seasonality(name='monthly', period=30.5, fourier_order=5)
            forecast = _prof.fit(df).predict(future)
            fig2 = _prof.plot_components(forecast)
            st.pyplot(fig2)
        
        @st.cache_data
        def ler_nomes_das_planilhas(caminho_arquivo):
            workbook = openpyxl.load_workbook(caminho_arquivo)
        
            # Obtém os nomes das planilhas
            nomes_planilhas = workbook.sheetnames
        
            # Fecha o arquivo Excel
            workbook.close()
            
            return nomes_planilhas
        
        @st.cache_data
        def generate_forecast_table(forecast, title):
            st.write(title)
            #st.write(forecast)
            forecast_renamed = forecast.rename(columns={'ds': 'Data', 'y': 'Previsão média'})
            return forecast_renamed


        @st.cache_data
        def calcular_porcentagem_entre_min_max(prev_min, prev_max, test_data):
            total = len(test_data)
            dentro_intervalo = ((test_data >= prev_min) & (test_data <= prev_max)).sum()
            percentual = dentro_intervalo / total * 100
            return percentual

        @st.cache_data
        def calcular_erros(test_data, forecast_train, st, option2):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                # Calcular o Erro Médio Absoluto (MAE)
                mae = mean_absolute_error(test_data['QUANT'], forecast_train)
                st.write("Erro Médio Absoluto (MAE):", int(mae))
            
            APE = mae / test_data['QUANT'] * 100
            
            # Calcular o Erro Percentual Absoluto Médio (MAPE)

            MAPE = APE.mean()
            
            # Configurar as cores para o gráfico de medidor
            if MAPE < 10:
                bar_color = "green"
            elif MAPE < 20:
                bar_color = "yellow"
            else:
                bar_color = "red"    
            
            # Criar o gráfico de medidor
            fig = go.Figure(go.Indicator(
                mode = "gauge+number",
                value = MAPE,
                title = {'text': "Erro Percentual Absoluto Médio (MAPE)"},
                gauge = {
                    'axis': {'range': [None, 100]},
                    'bar': {'color': bar_color, 'thickness': 1},  # Espessura máxima da barra
                    'bgcolor': 'black',
                    'borderwidth': 0,  # Remover bordas
                    'bordercolor': 'black',  # Ajustar a cor da borda para preto
                    'steps': [
                        {'range': [0, 100], 'color': 'gray'}  # Fundo preto
                    ],
                    'threshold': {
                        'line': {'color': "black", 'width': 4},
                        'thickness': 0.75,
                        'value': MAPE
                    }
                }
            ))
            with col2: 
                st.write("Erro Percentual Absoluto Médio (MAPE):", round(MAPE, 2), "%")
                #st.plotly_chart(fig)
            with col3:
                # Calcular o Root Mean Square Error (RMSE)
                RMSE = np.sqrt(mean_squared_error(test_data['QUANT'], forecast_train))
                st.write("Raiz do Erro Quadrático Médio (RMSE):", int(RMSE))

            
            # Calcular os erros
            errors = test_data['QUANT'] - forecast_train
        
            if option2:
                st.write("Diferenças entre os valores e as previsões:")
                st.write(errors)

            col1, col2 = st.columns(2)
              
            # Plotar o histograma dos erros
            plt.figure(figsize=(8, 6))
            plt.hist(errors, bins=6, color='skyblue', edgecolor='black')
            plt.title('Histograma dos Erros')
            plt.xlabel('Erro')
            plt.ylabel('Frequência')
            plt.grid(True)
            with col1: 
                st.pyplot(plt)
            
            # Erros em porcentagem
            errors_percent = (test_data['QUANT'] - forecast_train) / test_data['QUANT'] * 100
            if option2:
                st.write("Diferenças entre os valores e as previsões em porcentagem")
                st.write(errors_percent)
            
            # Plotar o histograma dos erros percentuais
            plt.figure(figsize=(8, 6))
            plt.hist(errors_percent, bins=6, color='skyblue', edgecolor='black')
            plt.title('Histograma dos Erros percentuais')
            plt.xlabel('Erro %')
            plt.ylabel('Frequência')
            plt.grid(True)

            with col2: 
                st.pyplot(plt)


# Obtenção dos dados da interface do streamlit
        st.title("Medidas de análise de precisão")
        with st.sidebar:
            ordem_filtro = st.slider("Ordem do filtro(semanas)", 1, 52)
            p = st.slider('Parâmetro p (AutoRegressivo)', 0, 20, 1)


            
            d = st.slider('Parâmetro d (Integração)', 0, 2, 1)
            q = st.slider('Parâmetro q (Média móvel)', 0, 25, 1)      

        col1, col2 = st.columns(2)

        ##grafico sem média móvel
        selected_productS = df
        selected_productS = moving_average_filter(selected_productS, 1)
        selected_product_titleS = selected_graficos + " - Ordem do Filtro: " + str(1) + " (semanas)"
        with col1:
            plot_historical_data(selected_productS, selected_product_titleS) 

        
        #primeiro gráfico
        
        #Utilização dos dados do streamlit para a execução
        selected_product = df
        selected_product = moving_average_filter(selected_product, ordem_filtro)
        selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        selected_product_title2 = "boxplot de " + selected_graficos + " - Ordem do Filtro: " + str(ordem_filtro) + " (semanas)"
        with col2:
            plot_historical_data(selected_product, selected_product_title) 
        
        
        #segundo gráfico
        #boxplot_historical_data(selected_product, selected_product_title2)

        #dividindo dados para treino e teste
        train_data, test_data, tamanhoPrevisao = train_test_split(selected_product)
        tamanhoPrevisao += 1

        
        #terceiro gráfico/previsão
        #periodo = st.slider("Intervalo a ser previsto(semanas)", 12, 38)      
        
        prof_train, forecast_train, df_train, future_train, fig, future_dates = create_profet_object(train_data, tamanhoPrevisao, p, d, q)
        #prof_train, forecast_train, df_train, future_train = create_profet_object(train_data, 38)
        forecast_train = forecast_train.iloc[0:]

        
        
        col1, col2, col3 = st.columns(3)
        with col1:
            #generate_forecast_report(prof_train, forecast_train, 'Previsão de demanda')
            st.pyplot(fig)
        with col2:   
            
            prof_test, forecast_test, df_test, future_test, fig, future_dates = create_profet_object(test_data, 1, p, d, q)

            # Criar o gráfico
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Plotar a linha de previsão
            ax.plot(test_data['DATA'], forecast_train, color='orange', linestyle='--', label='Previsão')
            
            # Plotar a linha de dados reais
            ax.plot(test_data['DATA'], test_data['QUANT'], color='blue', label='Real')
            
            # Adicionar títulos e legendas
            ax.set_title('Previsão vs Real')
            ax.set_xlabel('Data')
            ax.set_ylabel('Quantidade')
            ax.legend()
            
            # Exibir o gráfico no Streamlit
            st.pyplot(fig)
            
        with col3:
            #dividindo dados para treino e teste
            train_dataS, test_dataS, tamanhoPrevisao = train_test_split(selected_productS)
            prof_testS, forecast_testS, df_testS, future_testS, figS, future_datesS = create_profet_object(test_dataS, 1, p, d, q)
            
           # Criar o gráfico
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Plotar a linha de previsão
            ax.plot(test_data['DATA'], forecast_train, color='orange', linestyle='--', label='Previsão')
            
            # Plotar a linha de dados reais
            ax.plot(test_data['DATA'], test_dataS['QUANT'], color='blue', label='Real')
            
            # Adicionar títulos e legendas
            ax.set_title('Previsão vs Real (sem média móvel)')
            ax.set_xlabel('Data')
            ax.set_ylabel('Quantidade')
            ax.legend()
            
            # Exibir o gráfico no Streamlit
            st.pyplot(fig)
            
        option2 = st.checkbox('Mostrar detalhes dos gráficos')
        #option2 = 'Mostrar detalhes dos gráficos'
        #tabela previsão
        if option2:
            st.write(forecast_train)
        
    
        # Criar uma caixa de seleção
        #option = st.checkbox('Mostrar dados de análise de precisão')
        option = 'Mostrar dados de análise de precisão'
    
        
        # Verificar se a caixa de seleção está marcada
        
        if option:
            #if periodo != 0:
            #    forecast_table = forecast_table[:-periodo]
            #    test_data = test_data[:-periodo]
    
            # Redefinir os índices para garantir que estejam alinhados corretamente
            test_data.reset_index(drop=True, inplace=True)
            forecast_train.reset_index(drop=True, inplace=True)
            
            if option2:
                st.write("Dados para teste:")
                st.write(test_data)     
                st.write("Coluna previsão média")
                
                #st.write(forecast_train['predicted_mean']
                st.write(forecast_train)
                
                st.write("Coluna dados para teste")
                st.write(test_data['QUANT'])

            option3 = st.checkbox('Mostrar em relação aos dados sem média móvel')
            if option3:
                ############ Em relação a sem média móvel
                st.write('Em relação ao dado sem média móvel')
                selected_product = df
                selected_product = moving_average_filter(selected_product, 1)
                selected_product_title = selected_graficos + " - Ordem do Filtro: " + str(1) + " (semanas)"
                train_data, test_data, tamanhoPrevisao = train_test_split(selected_product)
                test_data.reset_index(drop=True, inplace=True)
                
                if option2:
                    st.write("Dados para teste-sem média móvel:")
                    st.write(test_data)
                
            ####função
            calcular_erros(test_data, forecast_train, st, option2)

def Pycaret(data):
    data = data.sort_values('DATA')
    data = data.set_index('DATA')

    # Cria índice semanal contínuo
    full_index = pd.date_range(start=data.index.min(), end=data.index.max(), freq='W-SUN')
    data = data.reindex(full_index)
    
    # Interpola valores ausentes
    data['QUANT'] = data['QUANT'].interpolate()

    # Prepara para o PyCaret
    data.index.name = 'DATA'
    data = data.reset_index()
    data = data[['DATA', 'QUANT']]

    # Inicia experimento
    exp = TSForecastingExperiment()
    exp.setup(data=data, target='QUANT', index='DATA', fh=3, session_id=123)

    # Compara modelos
    best = exp.compare_models()

    # Pega e exibe a tabela com as métricas dos modelos
    results = exp.pull()
    print(results)

    # Plota o gráfico da série temporal com previsão do melhor modelo
    exp.plot_model(best, plot='forecast')

        

def run_main_program():
    pd.options.mode.chained_assignment = None
    ######### caminho direto ou escolha
    #name = "D:\OneDrive\Área de Trabalho\CantoDeMinas\Dados semanais com gráficos.xlsx"

    menu = '0'
    with st.sidebar:
        database = st.checkbox('Utilizar banco de dados')
        
    if database:
        user = 'luan'
        password = 'luan'
        host = 'localhost'
        database = 'cantodeminas'
        connection = connect_to_mysql(user, password, host, database)

        
        # Obter nomes das tabelas
        name = get_table_names(connection)
        
        # Usar os nomes das tabelas para seleção
        with st.sidebar:
            selected_graficos = st.selectbox("Tabela:", name)
        df = read_table(connection, selected_graficos)
                       
        menu = '1'
        if connection is None:
            st.error("Não foi possível conectar ao banco de dados MySQL. Verifique as credenciais.")
            return
    if not database:
        name = st.file_uploader("Escolha o arquivo Excel", type=['xlsx'])
        #name = "D:\OneDrive\Área de Trabalho\CantoDeMinas\Dados semanais com gráficos.xlsx"

    #verificar se possui um arquivo presente
    if name is not None:
        # Criação do menu lateral
        # Obtenção dos dados da interface do streamlit
        if not database:
            with st.sidebar:
                selected_graficos = st.selectbox("Produtos:", ler_nomes_das_planilhas(name))
                
            df = read_sheet(name, selected_graficos)
        
            
            
            menu = '1'
        # Sliders para escolher a porcentagem de dados a remover
        remove_start = st.slider("Apagar dados do início (%)", 0, 100, 0)
        remove_end = st.slider("Apagar dados do fim (%)", 0, 100, 0)

        # Convertendo porcentagens para número de linhas
        rows_to_remove_start = int(len(df) * (remove_start / 100))
        rows_to_remove_end = int(len(df) * (remove_end / 100))

        # Garantindo que a remoção não exceda o total de linhas disponíveis
        if rows_to_remove_start + rows_to_remove_end >= len(df):
            st.warning("A remoção total excede o número de linhas disponíveis. Ajuste os sliders.")
            df_filtered = pd.DataFrame()  # Retorna um DataFrame vazio
        else:
            df_filtered = df.iloc[rows_to_remove_start: len(df) - rows_to_remove_end].copy()
        
    if menu=='1':
        with st.sidebar:
            selecao = option_menu(
                "Menu",
                ["Boxplot", "Previsão Prophet", "Análise Prophet", "Previsão NeuralProphet", "Análise NeuralProphet", "Previsão Arima", "Análise Arima", "Pycaret"],
                icons=['box', 'graph-up', 'bar-chart-line', 'graph-up', 'bar-chart-line', 'graph-up', 'bar-chart-line', 'bar-chart-line'],
                menu_icon="cast",
                default_index=0,
            )
        
        
        # Conteúdo da página principal baseado na seleção do menu
        if selecao == 'Previsão Prophet':
            previsao(df_filtered, name, selected_graficos)
        elif selecao == 'Boxplot':
            boxplot(df_filtered, name, selected_graficos)
        elif selecao == 'Análise Prophet':
            analise(df_filtered, name, selected_graficos)
        elif selecao == 'Previsão NeuralProphet':
            previsaoNeural(df_filtered, name, selected_graficos)
        elif selecao == 'Análise NeuralProphet':
            analiseNeural(df_filtered, name, selected_graficos)
        elif selecao == 'Previsão Arima':
            previsaoArima(df_filtered, name, selected_graficos)
        elif selecao == 'Análise Arima':
            analiseArima(df_filtered, name, selected_graficos)
        elif selecao == 'Pycaret':
            (df_filtered)
            
            
if __name__ == "__main__":
    main()
